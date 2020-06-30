from openpyxl import Workbook
from collections import namedtuple

# create workbook and sheet, fields
work_book = Workbook()
work_book.create_sheet("Sheet_A")

work_sheet = work_book.active
work_sheet.title = "Test"
work_sheet['A1'] = "Characters"
work_sheet['B1'] = "Numbers"

# generating test data
data = namedtuple('data', 'character number')
some_data = [data(character=c, number=n) for c, n in zip('abcde', (1, 2, 3, 4, 5))]

# inserting data to sheet
for index, data in enumerate(some_data):
    work_sheet.cell(index + 2, 1).value = data.character
    work_sheet.cell(index + 2, 2).value = data.number

# printing data
print(f'Sheet: {work_sheet.title}')
for row in work_sheet:
    a, b = row
    print(str(a.value).ljust(12), str(b.value).ljust(22))
